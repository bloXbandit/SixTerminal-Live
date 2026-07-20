# -*- coding: utf-8 -*-
"""
importer.py — Deterministic schedule extraction from Excel and PDF.

Pulls activities, WBS structure, durations, dates and status out of a schedule
export and normalizes them into a single source-agnostic "extraction contract"
(the same wbs[] / activities[] / relations[] backbone the LLM create_project
path uses, plus a per-row `_review` block for the confirm UI).

DESIGN PRINCIPLES
  • Deterministic first, offline by default. Excel (openpyxl) and text-layer
    PDF (pdfplumber / optional Tabula) never touch a network or a model — no
    schedule data leaves the machine unless the caller explicitly opts in.
  • The LLM is an OPTIONAL assist (use_llm=True), only reached for scanned/photo
    PDFs with no text layer, and only when a key is configured.
  • Robust column mapping + row classification so standard P6 / MSP exports come
    through with few or no bad pulls; anything uncertain is flagged for review
    rather than silently guessed.

Public API
  extract(path, source_type=None, pdf_engine="auto", use_llm=False) -> contract
  build_project_from_contract(contract, project_id=None) -> engine.schedule_model.Project
"""

import os
import re
import datetime as _dt
from typing import List, Dict, Any, Optional, Tuple


# ──────────────────────────────────────────────────────────────────────────────
# Column vocabulary — synonyms for the fields we care about (lowercased, no punct)
# ──────────────────────────────────────────────────────────────────────────────
_COLUMN_SYNONYMS: Dict[str, List[str]] = {
    "activity_id": [
        "activity id", "act id", "activityid", "task id", "id", "activity code",
        "activity", "task", "unique id",
    ],
    "name": [
        "activity name", "task name", "activity description", "description",
        "name", "title",
    ],
    "duration": [
        "at completion duration", "original duration", "remaining duration",
        "duration", "dur", "od", "rd", "planned duration",
    ],
    "start": [
        "start", "start date", "early start", "actual start", "planned start",
        "bl project start", "es",
    ],
    "finish": [
        "finish", "finish date", "end", "end date", "early finish",
        "actual finish", "planned finish", "ef",
    ],
    "wbs": ["wbs", "wbs code", "wbs name", "wbs path"],
    "predecessors": ["predecessors", "predecessor", "pred", "preds"],
    "successors": ["successors", "successor", "succ", "succs"],
    "percent": [
        "% complete", "percent complete", "activity % complete", "% comp",
        "pct complete", "complete",
    ],
    "type": ["activity type", "type"],
}

# An activity ID looks alphanumeric and contains at least one digit
# (e.g. A1000, MDC1.MIL.1000, EC-100). A WBS/group header does not.
_ID_PATTERN = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._\-/]*\d[A-Za-z0-9._\-/]*$")

# Name verbs that mark a zero-duration row as a milestone, and its direction
_FINISH_VERBS = (
    "finish", "complete", "completion", "received", "receive", "executed",
    "execute", "ready", "available", "issued", "acceptance", "occupancy",
    "energization", "energized", "released", "release", "tco", "closeout",
    "substantial",
)
_START_VERBS = ("start", "begin", "kickoff", "kick off", "commence", "ntp",
                "notice to proceed", "mobilize", "mobilization")

_DATE_FORMATS = (
    "%d-%b-%y", "%d-%b-%Y", "%d %b %y", "%d %b %Y",
    "%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d", "%b %d, %Y", "%d-%b",
)


def _norm(s: Any) -> str:
    """Lowercase, collapse whitespace, drop punctuation — for header matching."""
    if s is None:
        return ""
    t = re.sub(r"[^a-z0-9%\s]", " ", str(s).lower())
    return re.sub(r"\s+", " ", t).strip()


# ──────────────────────────────────────────────────────────────────────────────
# Value normalizers
# ──────────────────────────────────────────────────────────────────────────────
def _parse_date(raw: Any) -> Tuple[Optional[str], bool, bool]:
    """
    Return (iso_date | None, is_actual, parsed_ok).
    Strips the P6 actual suffix 'A' and the constraint marker '*'.
    """
    if raw is None:
        return None, False, True  # empty is fine, not a failure
    if isinstance(raw, (_dt.datetime, _dt.date)):
        d = raw.date() if isinstance(raw, _dt.datetime) else raw
        return d.isoformat(), False, True
    s = str(raw).strip()
    if not s or s in ("-", "--"):
        return None, False, True
    is_actual = False
    # trailing " A" (actual) — token, not part of a month like "Mar"
    m = re.search(r"\bA\b\s*$", s)
    if m:
        is_actual = True
        s = s[: m.start()].strip()
    s = s.rstrip("*").strip()
    for fmt in _DATE_FORMATS:
        try:
            d = _dt.datetime.strptime(s, fmt).date()
            return d.isoformat(), is_actual, True
        except ValueError:
            continue
    return None, is_actual, False  # had text but couldn't parse → flag it


def _parse_duration(raw: Any) -> Tuple[float, bool]:
    """Return (days, parsed_ok). Non-numeric / blank → 0."""
    if raw is None or raw == "":
        return 0.0, True
    if isinstance(raw, (int, float)):
        return float(raw), True
    s = str(raw).strip().lower()
    if not s:
        return 0.0, True
    weeks = "w" in s or "week" in s
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return 0.0, False
    val = float(m.group(0))
    if weeks:
        val *= 5.0  # working days per week
    return val, True


def _looks_like_id(v: Any) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    return bool(s) and bool(_ID_PATTERN.match(s)) and not s.replace(".", "").isalpha()


def _wbs_code_from_id(activity_id: str) -> Optional[str]:
    """MDC1.MIL.1000 -> MDC1.MIL ; A1000 -> None (no dotted prefix)."""
    if "." in activity_id:
        parts = activity_id.rsplit(".", 1)
        return parts[0] if parts[0] else None
    return None


def _infer_type(name: str, duration_days: float, explicit: Optional[str]) -> Tuple[str, bool]:
    """Return (activity_type, guessed). Honors an explicit type column first."""
    if explicit:
        e = _norm(explicit)
        table = {
            "task dependent": "Task Dependent", "task": "Task Dependent",
            "resource dependent": "Resource Dependent",
            "level of effort": "Level of Effort", "loe": "Level of Effort",
            "wbs summary": "WBS Summary",
            "start milestone": "Start Milestone",
            "finish milestone": "Finish Milestone",
            "milestone": "Finish Milestone",
        }
        for k, v in table.items():
            if k in e:
                return v, False
    if duration_days == 0:
        low = name.lower()
        if any(v in low for v in _START_VERBS):
            return "Start Milestone", True
        if any(v in low for v in _FINISH_VERBS):
            return "Finish Milestone", True
        return "Finish Milestone", True  # zero-duration default
    return "Task Dependent", False


# ──────────────────────────────────────────────────────────────────────────────
# Header detection + column mapping
# ──────────────────────────────────────────────────────────────────────────────
def _map_columns(rows: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    """
    Find the header row within the first 15 rows and map canonical field -> col idx.
    Returns (header_row_index, mapping). header_row_index is -1 if none found.
    """
    best_idx, best_map, best_score = -1, {}, 0
    for r_idx in range(min(15, len(rows))):
        cells = rows[r_idx]
        mapping: Dict[str, int] = {}
        for c_idx, cell in enumerate(cells):
            key = _norm(cell)
            if not key:
                continue
            for field, syns in _COLUMN_SYNONYMS.items():
                if field in mapping:
                    continue
                # exact match first, then whole-word contains
                if key in syns or any(re.search(rf"\b{re.escape(s)}\b", key) for s in syns):
                    mapping[field] = c_idx
                    break
        score = len(mapping) + (2 if "activity_id" in mapping else 0) \
            + (2 if "name" in mapping else 0)
        if score > best_score:
            best_idx, best_map, best_score = r_idx, mapping, score
    # Require at least id/name (or name + a date/duration) to trust it
    if "name" in best_map and ("activity_id" in best_map or "duration" in best_map
                               or "start" in best_map or "finish" in best_map):
        return best_idx, best_map
    return -1, {}


def _is_header_row(row: List[Any]) -> bool:
    """
    True if this row looks like a repeated column-header band. Multi-page PDF
    exports reprint the header on every page; without this they'd be mistaken
    for WBS section rows.
    """
    hits = 0
    for cell in row:
        key = _norm(cell)
        if not key:
            continue
        for syns in _COLUMN_SYNONYMS.values():
            if key in syns or any(re.search(rf"\b{re.escape(s)}\b", key) for s in syns):
                hits += 1
                break
    return hits >= 2


def _is_page_furniture(text: Any) -> bool:
    """Page numbers / lone date stamps printed in headers and footers."""
    if text is None:
        return False
    s = str(text).strip()
    if not s:
        return False
    if re.fullmatch(r"(?i)page\s+\d+(\s+of\s+\d+)?", s):
        return True
    iso, _, ok = _parse_date(s)
    return bool(ok and iso)          # a bare date line is furniture, not a WBS


def _backfill_columns(rows: List[List[Any]], header_idx: int,
                      cols: Dict[str, int]) -> Dict[str, int]:
    """
    Robustness net: if the header row missed a column (truncated/odd header),
    recover it from the DATA — a numeric-heavy column is the duration, and
    date-heavy columns are start/finish. Prevents silent 0-duration / blank-date
    pulls when a header label doesn't match the synonym list.
    """
    data = rows[header_idx + 1:] if header_idx >= 0 else rows
    if not data:
        return cols
    ncols = max((len(r) for r in rows), default=0)
    mapped = set(cols.values())

    def profile(c: int):
        num = dates = nonempty = 0
        for r in data:
            if c >= len(r):
                continue
            v = r[c]
            s = "" if v is None else str(v).strip()
            if not s:
                continue
            nonempty += 1
            iso, _, ok = _parse_date(s)
            if ok and iso:
                dates += 1
            elif re.fullmatch(r"-?\d+(?:\.\d+)?", s):
                num += 1
        return num, dates, nonempty

    # duration = the unmapped numeric-heavy column
    if "duration" not in cols:
        best, best_n = None, 0
        for c in range(ncols):
            if c in mapped:
                continue
            num, dates, ne = profile(c)
            if ne and dates == 0 and num / ne > 0.6 and num > best_n:
                best, best_n = c, num
        if best is not None:
            cols["duration"] = best
            mapped.add(best)

    # start / finish = unmapped date-heavy columns, left-to-right
    date_cols = []
    for c in range(ncols):
        if c in mapped:
            continue
        num, dates, ne = profile(c)
        if ne and dates / ne > 0.5:
            date_cols.append(c)
    for field in ("start", "finish"):
        if field not in cols and date_cols:
            cols[field] = date_cols.pop(0)
            mapped.add(cols[field])
    return cols


# ──────────────────────────────────────────────────────────────────────────────
# Core: rows -> contract
# ──────────────────────────────────────────────────────────────────────────────
def _rows_to_contract(rows: List[List[Any]], meta: Dict[str, Any],
                      outline_levels: Optional[List[int]] = None) -> Dict[str, Any]:
    warnings: List[str] = []
    header_idx, cols = _map_columns(rows)
    if header_idx < 0:
        warnings.append("Could not identify a header row — using positional guess "
                        "(col1=ID, col2=Name).")
        cols = {"activity_id": 0, "name": 1}
        header_idx = -1
    cols = _backfill_columns(rows, header_idx, cols)

    def cell(row, field):
        idx = cols.get(field)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return v if v not in ("", None) else None

    wbs_nodes: List[Dict[str, Any]] = []
    wbs_by_code: Dict[str, Dict[str, Any]] = {}
    activities: List[Dict[str, Any]] = []
    relations: List[Dict[str, Any]] = []

    # header stack for hierarchy (outline level or prefix depth) -> current node
    header_stack: List[Tuple[int, Dict[str, Any]]] = []
    current_wbs_code: Optional[str] = None
    seen_ids = set()
    date_fail = 0
    logic_seen = False
    seq_counter = 0

    def ensure_wbs(code: str, name: str, parent_code: Optional[str],
                   confidence: float, derived: List[str], source: Dict) -> Dict:
        node = wbs_by_code.get(code)
        if node:
            if name and node["name"] in (code, ""):
                node["name"] = name
            return node
        nonlocal seq_counter
        node = {
            "code": code, "name": name or code, "parent_code": parent_code,
            "sequence": seq_counter,
            "_review": {"confidence": confidence, "source": source, "derived": derived},
        }
        seq_counter += 1
        wbs_nodes.append(node)
        wbs_by_code[code] = node
        return node

    def ensure_prefix_chain(prefix: str, src: Dict) -> str:
        """Flat-export fallback: build nested nodes from a dotted ID prefix
        (MDC1.MIL -> MDC1 > MDC1.MIL). Returns the leaf code."""
        parts = prefix.split(".")
        parent, code = None, prefix
        for i in range(1, len(parts) + 1):
            code = ".".join(parts[:i])
            if code not in wbs_by_code:
                ensure_wbs(code, code, parent, 0.7, ["code", "name"], src)
            parent = code
        return code

    for r_idx in range(header_idx + 1, len(rows)):
        row = rows[r_idx]
        if not any(v not in ("", None) for v in row):
            continue
        src = {"row": r_idx + 1}
        raw_id = cell(row, "activity_id")
        raw_name = cell(row, "name")
        # some exports put the WBS header text in the ID column, name blank
        display = raw_name if raw_name is not None else raw_id
        level = outline_levels[r_idx] if outline_levels and r_idx < len(outline_levels) else None

        # Multi-page hygiene: drop reprinted column headers and page furniture.
        # Activity rows always carry a valid ID, so they can never be dropped here.
        if not _looks_like_id(raw_id):
            if _is_header_row(row) or _is_page_furniture(display):
                continue

        if _looks_like_id(raw_id):
            # ── ACTIVITY ROW ──────────────────────────────────────────────────
            act_id = str(raw_id).strip()
            name = str(raw_name).strip() if raw_name else act_id
            if act_id in seen_ids:
                continue  # dedupe repeated header/page rows
            seen_ids.add(act_id)

            dur_raw = cell(row, "duration")
            dur_days, dur_ok = _parse_duration(dur_raw)

            s_iso, s_actual, s_ok = _parse_date(cell(row, "start"))
            f_iso, f_actual, f_ok = _parse_date(cell(row, "finish"))
            if not s_ok or not f_ok:
                date_fail += 1

            explicit_type = cell(row, "type")
            atype, type_guessed = _infer_type(name, dur_days, explicit_type)

            # status from actuals
            pct_raw = cell(row, "percent")
            pct = None
            if pct_raw is not None:
                pv, _ = _parse_duration(pct_raw)
                pct = max(0.0, min(100.0, pv))
            if s_actual and f_actual:
                status, pct = "Completed", 100.0
            elif s_actual:
                status = "In Progress"
                pct = pct if pct not in (None, 0.0) else 50.0
            else:
                status = "Not Started"
                pct = pct if pct is not None else 0.0

            # WBS assignment: the section header the row sits under is primary
            # (real names + hierarchy). Fall back to the activity-ID prefix only
            # when the export has no header rows at all (flat table).
            derived_fields = []
            flags = []
            if current_wbs_code:
                wbs_code = current_wbs_code
            else:
                id_wbs = _wbs_code_from_id(act_id)
                if id_wbs:
                    wbs_code = ensure_prefix_chain(id_wbs, src)
                    derived_fields.append("wbs_code")
                else:
                    wbs_code = None
                    flags.append("wbs_unresolved")

            if type_guessed:
                flags.append("type_guessed")
            if not dur_ok:
                flags.append("duration_unparsed")
            if (not s_ok) or (not f_ok):
                flags.append("date_unparsed")

            # relationships (usually absent in a printed/exported view)
            for fld, kind in (("predecessors", "pred"), ("successors", "succ")):
                rawlink = cell(row, fld)
                if rawlink:
                    logic_seen = True
                    for tok in re.split(r"[,\n;]+", str(rawlink)):
                        tok = tok.strip()
                        if not tok:
                            continue
                        mid = re.match(r"([A-Za-z0-9._\-/]+)", tok)
                        if not mid:
                            continue
                        other = mid.group(1)
                        rtype = "fs"
                        mt = re.search(r"\b(fs|ss|ff|sf)\b", tok.lower())
                        if mt:
                            rtype = mt.group(1)
                        if kind == "pred":
                            relations.append({"predecessor_id": other,
                                              "successor_id": act_id, "type": rtype})
                        else:
                            relations.append({"predecessor_id": act_id,
                                              "successor_id": other, "type": rtype})

            confidence = 1.0
            confidence -= 0.15 * len(flags)
            confidence -= 0.05 * len([d for d in derived_fields])
            confidence = round(max(0.3, confidence), 2)

            activities.append({
                "activity_id": act_id,
                "name": name,
                "wbs_code": wbs_code,
                "type": atype,
                "duration_days": round(dur_days, 2),
                "status": status,
                "percent_complete": round(pct or 0.0, 1),
                "planned_start": None if s_actual else s_iso,
                "planned_finish": None if f_actual else f_iso,
                "actual_start": s_iso if s_actual else None,
                "actual_finish": f_iso if f_actual else None,
                "constraint_type": None,
                "constraint_date": None,
                "_review": {
                    "confidence": confidence,
                    "source": src,
                    "raw": {"duration": _s(dur_raw), "start": _s(cell(row, "start")),
                            "finish": _s(cell(row, "finish"))},
                    "derived": derived_fields,
                    "flags": flags,
                },
            })
        elif display and str(display).strip():
            # ── WBS HEADER ROW ────────────────────────────────────────────────
            name = str(display).strip()
            if _norm(name) in ("total", "grand total"):
                continue
            depth = level if level is not None else (name.count(".") if "." in name else 0)
            # pop deeper/equal headers off the stack
            while header_stack and header_stack[-1][0] >= depth:
                header_stack.pop()
            # A band reprinted on a continuation page is the SAME section —
            # reuse the existing node instead of creating a duplicate.
            existing = next((n for n in wbs_nodes if _norm(n["name"]) == _norm(name)), None)
            if existing:
                node = existing
            else:
                parent_code = header_stack[-1][1]["code"] if header_stack else None
                node = ensure_wbs(_synth_wbs_code(name, wbs_by_code), name,
                                  parent_code, 0.85, [], src)
            header_stack.append((depth, node))
            current_wbs_code = node["code"]

    # ── finalize meta ────────────────────────────────────────────────────────
    if date_fail:
        warnings.append(f"{date_fail} date value(s) could not be parsed — check flagged rows.")
    if not activities:
        warnings.append("No activity rows were detected. Is this a schedule export?")
    logic_status = "complete" if (logic_seen and relations) else \
        ("partial" if relations else "absent")
    if logic_status == "absent":
        warnings.append("No predecessor/successor logic found — the agent can propose "
                        "ties after import.")

    meta = dict(meta)
    meta.update({
        "logic_status": logic_status,
        "warnings": warnings,
        "column_map": {k: v for k, v in cols.items()},
        "activity_count": len(activities),
        "wbs_count": len(wbs_nodes),
        "relation_count": len(relations),
        "extraction_confidence": round(
            sum(a["_review"]["confidence"] for a in activities) / len(activities), 2
        ) if activities else 0.0,
    })
    return {"meta": meta, "wbs": wbs_nodes, "activities": activities,
            "relations": relations}


def _s(v):
    return None if v is None else str(v)


def _synth_wbs_code(name: str, existing: Dict[str, Any]) -> str:
    """Deterministic short code from a WBS name, unique within the set."""
    base = re.sub(r"[^A-Za-z0-9]", "", name).upper()[:12] or "WBS"
    code = base
    n = 1
    while code in existing:
        n += 1
        code = f"{base}{n}"
    return code


# ──────────────────────────────────────────────────────────────────────────────
# Source readers
# ──────────────────────────────────────────────────────────────────────────────
def _read_xlsx(path: str) -> Tuple[List[List[Any]], List[int]]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    # pick the sheet with the most populated rows
    best_ws, best_rows = None, -1
    for ws in wb.worksheets:
        if ws.max_row and ws.max_row > best_rows:
            best_ws, best_rows = ws, ws.max_row
    ws = best_ws
    rows: List[List[Any]] = []
    outline: List[int] = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        rows.append(list(row))
        dim = ws.row_dimensions.get(r_idx)
        outline.append(int(getattr(dim, "outline_level", 0) or 0) if dim else 0)
    return rows, outline


def _read_pdf(path: str, engine: str = "auto") -> Tuple[List[List[Any]], str]:
    """
    Return (rows, engine_used). Deterministic, offline.
      engine: "auto" | "pdfplumber" | "tabula"
    """
    used = ""
    rows: List[List[Any]] = []

    def _pdfplumber() -> List[List[Any]]:
        import pdfplumber
        out: List[List[Any]] = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables() or []
                for tbl in tables:
                    for r in tbl:
                        out.append([("" if c is None else str(c).replace("\n", " ").strip())
                                    for c in r])
                if not tables:
                    # fall back to word clustering for borderless tables
                    words = page.extract_words(use_text_flow=True) or []
                    out.extend(_cluster_words(words))
        return out

    def _tabula() -> List[List[Any]]:
        import tabula
        dfs = tabula.read_pdf(path, pages="all", multiple_tables=True,
                              lattice=True, pandas_options={"header": None},
                              silent=True)
        if not any(len(d) for d in dfs):
            dfs = tabula.read_pdf(path, pages="all", multiple_tables=True,
                                  stream=True, pandas_options={"header": None},
                                  silent=True)
        out: List[List[Any]] = []
        for d in dfs:
            for _, r in d.iterrows():
                out.append(["" if _isnan(c) else str(c).strip() for c in r.tolist()])
        return out

    if engine == "tabula":
        rows, used = _tabula(), "tabula"
    elif engine == "pdfplumber":
        rows, used = _pdfplumber(), "pdfplumber"
    else:  # auto — pdfplumber first (pure python), tabula as a stronger fallback
        rows = _pdfplumber()
        used = "pdfplumber"
        if _too_sparse(rows):
            try:
                t_rows = _tabula()
                if len(t_rows) > len(rows):
                    rows, used = t_rows, "tabula"
            except Exception:
                pass  # Java/tabula unavailable — keep pdfplumber result
    return rows, used


def _cluster_words(words: List[Dict]) -> List[List[str]]:
    """Group extracted words into rows by y, then columns by x gaps."""
    if not words:
        return []
    rows: Dict[int, List[Dict]] = {}
    for w in words:
        key = round(w["top"] / 4.0)  # ~4px row bucket
        rows.setdefault(key, []).append(w)
    out: List[List[str]] = []
    for key in sorted(rows):
        line = sorted(rows[key], key=lambda w: w["x0"])
        cells, buf, last_x1 = [], [], None
        for w in line:
            if last_x1 is not None and (w["x0"] - last_x1) > 24:  # column gap
                cells.append(" ".join(buf)); buf = []
            buf.append(w["text"]); last_x1 = w["x1"]
        if buf:
            cells.append(" ".join(buf))
        out.append(cells)
    return out


def _too_sparse(rows: List[List[Any]]) -> bool:
    """Heuristic: pdfplumber found little/no tabular structure."""
    if len(rows) < 5:
        return True
    multi = sum(1 for r in rows if len([c for c in r if str(c).strip()]) >= 3)
    return multi < max(3, len(rows) * 0.3)


def _isnan(v) -> bool:
    try:
        return v != v  # NaN
    except Exception:
        return False


def _text_layer_present(path: str) -> bool:
    """True if the PDF has extractable text (digital), False if it's a scan/image."""
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages[:3]:
                if (page.extract_text() or "").strip():
                    return True
    except Exception:
        pass
    return False


# ──────────────────────────────────────────────────────────────────────────────
# Public entry point
# ──────────────────────────────────────────────────────────────────────────────
def extract(path: str, source_type: Optional[str] = None,
            pdf_engine: str = "auto", use_llm: bool = False,
            api_key: Optional[str] = None, model_key: Optional[str] = None) -> Dict[str, Any]:
    """
    Extract a schedule into the contract. Deterministic and offline unless
    use_llm=True is passed for a scanned PDF with no text layer.
    """
    ext = (source_type or os.path.splitext(path)[1].lstrip(".")).lower()
    base_meta = {
        "source_type": ext, "source_name": os.path.basename(path),
        "project_name": os.path.splitext(os.path.basename(path))[0],
        "project_id": None, "data_date": None, "hours_per_day": 8,
        "engine": None, "llm_used": False,
    }

    if ext in ("xlsx", "xlsm", "xls"):
        rows, outline = _read_xlsx(path)
        base_meta["engine"] = "openpyxl"
        return _rows_to_contract(rows, base_meta, outline_levels=outline)

    if ext == "pdf":
        has_text = _text_layer_present(path)
        if not has_text and use_llm:
            return _extract_pdf_llm(path, base_meta, api_key, model_key)
        rows, used = _read_pdf(path, engine=pdf_engine)
        base_meta["engine"] = used
        contract = _rows_to_contract(rows, base_meta)
        if not has_text and not contract["activities"]:
            contract["meta"]["warnings"].insert(
                0, "This PDF has no text layer (looks scanned/photographed). "
                   "Deterministic extraction found nothing — enable AI assist to read it.")
        return contract

    raise ValueError(f"Unsupported file type: '{ext}'. Use .xlsx or .pdf.")


def _extract_pdf_llm(path: str, base_meta: Dict[str, Any],
                     api_key: Optional[str], model_key: Optional[str]) -> Dict[str, Any]:
    """
    OPT-IN vision fallback for scanned PDFs. Sends the document to the model.
    Kept isolated so the deterministic path never imports or reaches it.
    """
    raise NotImplementedError(
        "AI-assisted extraction is not enabled in this build. The deterministic "
        "Excel and text-layer PDF paths run fully offline.")


# ──────────────────────────────────────────────────────────────────────────────
# Materialize the (reviewed) contract into a Project
# ──────────────────────────────────────────────────────────────────────────────
def build_project_from_contract(contract: Dict[str, Any],
                                project_id: Optional[str] = None):
    """Turn a confirmed extraction contract into a schedule_model.Project."""
    from engine.schedule_model import (Project, WBSNode, Activity, Relation,
                                        Calendar, compute_dates)

    meta = contract.get("meta", {})
    hpd = float(meta.get("hours_per_day") or 8)
    proj_id = (project_id or meta.get("project_id") or meta.get("project_name")
               or "IMPORT")[:12]
    proj_uid = str(abs(hash(proj_id)))[:8]

    project = Project(
        uid=proj_uid,
        name=meta.get("project_name") or proj_id,
        id=proj_id,
        data_date=meta.get("data_date"),
    )
    project.calendars = [Calendar(uid="1", name="Standard", hours_per_day=hpd)]

    # WBS
    wbs_by_code: Dict[str, WBSNode] = {}
    # first pass: create nodes
    for w in contract.get("wbs", []):
        code = str(w.get("code") or "").strip()
        if not code:
            continue
        uid = str(abs(hash(code + proj_uid)))[:8]
        node = WBSNode(uid=uid, name=str(w.get("name") or code), code=code,
                       parent_uid=None, sequence_num=int(w.get("sequence", 0) or 0))
        project.wbs_nodes.append(node)
        wbs_by_code[code] = node
    # second pass: wire parents
    for w in contract.get("wbs", []):
        code, parent = str(w.get("code") or ""), w.get("parent_code")
        if code in wbs_by_code and parent and parent in wbs_by_code:
            wbs_by_code[code].parent_uid = wbs_by_code[parent].uid
    if not project.wbs_nodes:
        root = WBSNode(uid="10", name=project.name, code="ROOT")
        project.wbs_nodes.append(root)
        wbs_by_code["ROOT"] = root
    default_wbs_uid = project.wbs_nodes[0].uid

    # Activities
    act_by_id: Dict[str, Activity] = {}
    for a in contract.get("activities", []):
        act_id = str(a.get("activity_id") or "").strip()
        if not act_id or act_id in act_by_id:
            continue
        wbs_code = a.get("wbs_code")
        wbs_uid = wbs_by_code[wbs_code].uid if wbs_code in wbs_by_code else default_wbs_uid
        dur_h = float(a.get("duration_days") or 0) * hpd
        status = a.get("status") or "Not Started"
        act = Activity(
            uid=str(abs(hash(act_id + proj_uid)))[:8],
            activity_id=act_id,
            name=str(a.get("name") or act_id),
            wbs_uid=wbs_uid,
            calendar_uid="1",
            activity_type=a.get("type") or "Task Dependent",
            status=status,
            planned_duration=dur_h,
            remaining_duration=0.0 if status == "Completed" else dur_h,
            percent_complete=float(a.get("percent_complete") or 0),
            planned_start=a.get("planned_start"),
            planned_finish=a.get("planned_finish"),
            actual_start=a.get("actual_start"),
            actual_finish=a.get("actual_finish"),
        )
        project.activities.append(act)
        act_by_id[act_id] = act

    # Relations (only those whose endpoints both exist)
    rel_type_map = {"fs": "Finish to Start", "ss": "Start to Start",
                    "ff": "Finish to Finish", "sf": "Start to Finish"}
    for r in contract.get("relations", []):
        p, s = str(r.get("predecessor_id") or ""), str(r.get("successor_id") or "")
        if p in act_by_id and s in act_by_id and p != s:
            project.relations.append(Relation(
                uid=str(abs(hash(p + s + proj_uid)))[:8],
                predecessor_uid=act_by_id[p].uid,
                successor_uid=act_by_id[s].uid,
                type=rel_type_map.get(str(r.get("type", "fs")).lower(), "Finish to Start"),
                lag=float(r.get("lag_days", 0) or 0) * hpd,
            ))

    project.build_lookups()
    try:
        compute_dates(project)
    except Exception:
        pass
    return project
